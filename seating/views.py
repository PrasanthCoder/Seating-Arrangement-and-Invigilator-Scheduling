from django.shortcuts import render
import pandas as pd
import xlsxwriter as xlsxwriter
import openpyxl
from django.http import HttpResponse
from .models import *
from django.db.models import Count

# Create your views here.
def input(request):
        if request.method == 'POST':
            print("entered post")
            xcel_file1 = request.FILES['students']
            xcel_file2 = request.FILES['subjects']
            stu_file = pd.read_excel(xcel_file1)
            sub_file = pd.read_excel(xcel_file2)
            print("Entering loop")
            for index, row in stu_file.iterrows():
                print("Entered loop")
                if BTStudentInfo.objects.filter(RegNo=row['RegNo']).exists():
                    print("not working")
                    pass
                else:
                    print("working")
                    student_instance = BTStudentInfo(
                        RegNo = row['RegNo'],
                        RollNo = row['RollNo'],
                        Name = row['Name'],
                        Regulation = row['Regulation'],
                        Dept = row['Dept'],
                        AdmissionYear = row['AdmissionYear'],
                        Gender = row['Gender'],
                        Category = row['Category'],
                        GuardinaName = row['GuardianName'],
                        Phone = row['Phone'],
                        Email = row['Email'],
                        Address1 = row['Address1'],
                        Address2 = row['Address2'],
                        Cycle = row['Cycle']
                    )
                    student_instance.save()

                    roll_instance = BTRollLists(
                        Cycle = row['Cycle'],
                        Section = row['Section'],
                        Student = student_instance
                    )
                    roll_instance.save()

            for index, row in sub_file.iterrows():
                if BTSubjectInfo.objects.filter(SubName = row['SubName']).exists():
                    print("not working")
                    pass
                else:
                    print("working")
                    subject_instance = BTSubjectInfo(
                        Year = row['Year'],
                        Sem = row['Sem'],
                        Regulation = row['Regulation'],
                        Mode = row['Mode'],
                        Dept = row['Dept'],
                        SubId = row['SubId'],
                        SubCode = row['SubCode'],
                        SubName = row['SubName'],
                        Credits = row['Credits'],
                        Type = row['Type'],
                        Category = row['Category']
                    )
                    subject_instance.save()

            return render(request, 'input.html')
        else:
            return render(request, 'input.html')
     
def home(request):
    
    if request.method == 'POST':
        print(request.FILES['seats'])
        xcel_file1 = request.FILES['registrations']
        xcel_file2 = request.FILES['seats']
        reg_file = pd.read_excel(xcel_file1) # reading registrations
        seat_file = openpyxl.load_workbook(xcel_file2) # reading seating blocks
        

        writer = pd.ExcelWriter('output_seating.xlsx', engine='xlsxwriter') # creating an excel file
        print(reg_file)
        for index,row in reg_file.iterrows(): # entering registation info into the databse.
             Studentinfo = BTRollLists.objects.get(Student = BTStudentInfo.objects.get(RollNo = row['RollNo']))
             Subinfo = BTSubjectInfo.objects.get(SubCode = row['SubCode'])
             if BTStudnetRegistrations.objects.filter(Student=Studentinfo).exists():
                 pass
             else:
                reg_instance = BTStudnetRegistrations(
                    Student = Studentinfo,
                    Mode = row['Mode'],
                    Sub_Id = Subinfo
                )
                reg_instance.save()
        sub_list = list(BTStudnetRegistrations.objects.values('Sub_Id_id').annotate(sub_count=Count('Sub_Id_id')).order_by('-sub_count', 'Sub_Id_id').values_list('Sub_Id_id', flat=True).distinct())
        sub_len = len(sub_list)
        sub_no = 0
        sub1 = sub2 = sub3 = sub4 = []
        curr_subs = [sub1,sub2,sub3,sub4]
        if(sub_no < sub_len):
            curr_subs[0] = list(BTStudnetRegistrations.objects.filter(Sub_Id_id=sub_list[sub_no]).values_list('Student', flat=True))
            sub_no  += 1
        if(sub_no < sub_len):
            curr_subs[1] = list(BTStudnetRegistrations.objects.filter(Sub_Id_id=sub_list[sub_no]).values_list('Student', flat=True))
            sub_no  += 1
        if(sub_no < sub_len):  
            curr_subs[2] = list(BTStudnetRegistrations.objects.filter(Sub_Id_id=sub_list[sub_no]).values_list('Student', flat=True))
            sub_no  += 1
        if(sub_no < sub_len):
            curr_subs[3] = list(BTStudnetRegistrations.objects.filter(Sub_Id_id=sub_list[sub_no]).values_list('Student', flat=True))
            sub_no  += 1
        count = 1 # for sheet number
        curr_sub_no = 0
        print(curr_subs)

        for sheet_name in seat_file.sheetnames:
            worksheet = seat_file[sheet_name]
            df = pd.DataFrame() # creating dataframe
            for col in worksheet.iter_cols():
                curr_subs[0],curr_subs[2] = curr_subs[2],curr_subs[0]
                curr_subs[1],curr_subs[3] = curr_subs[3],curr_subs[1]
                curr_sub_no = 0
                for cell in col:
                    color = cell.fill.start_color.index
                    if color == 'FFFFFF00': # yellow colour
                        if(len(curr_subs[curr_sub_no]) != 0):
                            student = BTStudentInfo.objects.get(id=curr_subs[curr_sub_no][0])
                            df.loc[cell.row, cell.column] = student.RollNo
                            curr_subs[curr_sub_no].pop(0)
                            if(curr_sub_no>=3):
                                curr_sub_no = 0
                            else:
                                curr_sub_no += 1
                        else:
                            if(sub_no >= sub_len):
                                if(curr_sub_no>=3):
                                    curr_sub_no = 0
                                else:
                                    curr_sub_no += 1
                            else:
                                curr_subs[curr_sub_no] = list(BTStudnetRegistrations.objects.filter(Sub_Id_id=sub_list[sub_no]).values_list('Student', flat=True))
                                sub_no  += 1
                                student = BTStudentInfo.objects.get(id=curr_subs[curr_sub_no][0])
                                df.loc[cell.row, cell.column] = student.RollNo
                                curr_subs[curr_sub_no].pop(0)
                                if(curr_sub_no>=3):
                                    curr_sub_no = 0
                                else:
                                    curr_sub_no += 1
                    elif color == 'FFFF0000': # red color
                        df.loc[cell.row, cell.column] = "NA"
                        if(curr_sub_no>=3):
                            curr_sub_no = 0
                        else:
                            curr_sub_no += 1
                    else: 
                        continue

            pattern1 = 'C{}'
            pattern2 = 'R{}'

            col_names = [pattern1.format(i) for i in range(1, df.shape[1]+1)]
            row_names = [pattern2.format(i) for i in range(1, df.shape[0]+1)]

            df.index = row_names
            df.columns = col_names
            df.to_excel(writer, sheet_name='NewSheet'+str(count))
            print(df)
            count += 1

        writer.save()
        
        with open('output_seating.xlsx', 'rb') as f:
            buffer = f.read()
        
        response = HttpResponse(buffer, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="output_seating.xlsx"'
        return response
    else:
        return render(request, 'home.html')
    
